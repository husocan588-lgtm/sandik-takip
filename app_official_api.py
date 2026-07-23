from flask import Flask, request, jsonify, send_file, render_template, session, redirect, url_for
from werkzeug.security import check_password_hash
import io
import openpyxl
from flask_cors import CORS
from database import get_db_connection, is_point_in_polygon
import psycopg2
import json
import threading
import time
import requests
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'super-secret-sandik-takip-key'
CORS(app)

@app.before_request
def log_request_info():
    print(f"{request.method} {request.url}")

# Web Arayüzü Ana Sayfası
@app.route('/', methods=['GET'])
def index():
    return redirect(url_for('dashboard'))

# Web Oturum Açma
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()
        conn.close()
        
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error='Hatalı kullanıcı adı veya şifre')
            
    return render_template('login.html')

# Web Çıkış Yapma
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# Web Yönetim Paneli
@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('dashboard.html', username=session.get('username'))

# Web AOG Yönetimi
@app.route('/aog')
def aog():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('aog.html', username=session.get('username'))

# Yeni Sandık/Tag Kaydetme
@app.route('/api/crates', methods=['POST'])
def create_crate():
    data = request.json
    tag_mac = data.get('tag_mac')
    name = data.get('name')
    width = data.get('width')
    height = data.get('height')
    length = data.get('length')
    public_key = data.get('public_key')
    pn = data.get('pn')
    
    if not tag_mac or not name:
        return jsonify({'error': 'tag_mac ve name zorunludur.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            '''INSERT INTO crates (tag_mac, name, width, height, length, public_key, pn) VALUES (%s, %s, %s, %s, %s, %s, %s)''',
            (tag_mac, name, width, height, length, public_key, pn)
        )
        conn.commit()
        last_id = cursor.lastrowid
        return jsonify({'message': 'Sandık kaydedildi.', 'id': last_id}), 201
    except psycopg2.IntegrityError:
        return jsonify({'error': 'Bu Tag MAC adresi zaten kayıtlı.'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# Sandıkları Listeleme
@app.route('/api/crates', methods=['GET'])
def get_crates():
    sync_hbbk_data()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM crates')
    rows = cursor.fetchall()
    conn.close()
    
    crates = [dict(row) for row in rows]
                
    return jsonify(crates)

# Sandık Güncelleme (Düzenleme)
@app.route('/api/crates/<int:id>', methods=['PUT'])
def update_crate(id):
    data = request.json
    tag_mac = data.get('tag_mac')
    name = data.get('name')
    width = data.get('width')
    height = data.get('height')
    length = data.get('length')
    public_key = data.get('public_key')
    pn = data.get('pn')
    
    if not tag_mac or not name:
        return jsonify({'error': 'tag_mac ve name zorunludur.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            '''UPDATE crates SET tag_mac = %s, name = %s, width = %s, height = %s, length = %s, public_key = %s, pn = %s WHERE id = %s''',
            (tag_mac, name, width, height, length, public_key, pn, id)
        )
        conn.commit()
        if cursor.rowcount == 0:
            return jsonify({'error': 'Güncellenecek sandık bulunamadı.'}), 404
        return jsonify({'message': 'Sandık güncellendi.'})
    except psycopg2.IntegrityError:
        return jsonify({'error': 'Bu Tag MAC adresi zaten kayıtlı.'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# Sandık Silme
@app.route('/api/crates/<int:id>', methods=['DELETE'])
def delete_crate(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('DELETE FROM crates WHERE id = %s', (id,))
        conn.commit()
        if cursor.rowcount == 0:
            return jsonify({'error': 'Silinecek sandık bulunamadı.'}), 404
        return jsonify({'message': 'Sandık silindi.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# Geofence Koordinatlarını Getirme
@app.route('/api/geofence', methods=['GET'])
def get_geofence():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM settings WHERE key = 'geofence_polygons'")
    row = cursor.fetchone()
    conn.close()
    
    if row and row['value']:
        return jsonify(json.loads(row['value']))
    else:
        return jsonify([])

# Ping (Tarayıcı cihazlardan gelen konum verisi)
@app.route('/api/ping', methods=['POST'])
def ping_crate():
    data = request.json
    tag_mac = data.get('tag_mac')
    lat = data.get('lat')
    lng = data.get('lng')
    
    if not tag_mac or lat is None or lng is None:
        return jsonify({'error': 'tag_mac, lat ve lng zorunludur.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Yeşil Alan Koordinatlarını (Poligon Dizisini) alalım
        cursor.execute("SELECT value FROM settings WHERE key = 'geofence_polygons'")
        row = cursor.fetchone()
        
        status = 'Stokta Yok'
        if row and row['value']:
            polygons = json.loads(row['value'])
            # Gelen konum herhangi bir poligonun içindeyse Stokta Var sayalım
            for polygon in polygons:
                if is_point_in_polygon({'lat': lat, 'lng': lng}, polygon):
                    status = 'Stokta Var'
                    break
                
        # Sandığın konumunu ve durumunu güncelle
        cursor.execute(
            '''UPDATE crates SET last_seen_lat = %s, last_seen_lng = %s, last_seen_time = CURRENT_TIMESTAMP, status = %s WHERE tag_mac = %s''',
            (lat, lng, status, tag_mac)
        )
        conn.commit()
        
        return jsonify({'message': 'Ping alındı ve durum güncellendi.', 'status': status})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

def sync_hbbk_data():
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Resmi API Bilgileri
        api_key = "hk_lZZ8t1fifpKfRuMsrcZb1AVSz161I_Ti"
        base_url = "https://api.hbb-k.com/api/v1/devices"
        headers = {"X-API-KEY": api_key}
        
        # 1. Cihaz listesini al
        response = requests.get(base_url, headers=headers, timeout=10)
        if response.status_code != 200:
            print(f"[HBB-K Sync] API Hatası: {response.status_code}")
            return
            
        data = response.json().get("data", [])
        
        for item in data:
            pub_key = item.get("publicKey")
            mac = item.get("mac", pub_key)
            name = item.get("name") or item.get("deviceName", "API Cihazı")
            battery = item.get("latestBattery") or 0
            
            if not pub_key:
                continue
                
            lat = item.get("latestLat")
            lng = item.get("latestLng")
            timestamp = item.get("latestTime") or datetime.now().isoformat()
            
            if lat is not None and lng is not None:
                # Yeşil Alan (Geofence) Kontrolü
                cursor.execute("SELECT value FROM settings WHERE key = 'geofence_polygons'")
                poly_row = cursor.fetchone()
                
                status = 'Stokta Yok'
                if poly_row and poly_row['value']:
                    polygons = json.loads(poly_row['value'])
                    for polygon in polygons:
                        if is_point_in_polygon({'lat': lat, 'lng': lng}, polygon):
                            status = 'Stokta Var'
                            break
                
                # Veritabanında kontrol
                cursor.execute('SELECT id, last_seen_lat, last_seen_lng, last_seen_address FROM crates WHERE public_key = %s OR tag_mac = %s', (pub_key, mac))
                existing = cursor.fetchone()
                
                address = existing['last_seen_address'] if existing else None
                
                if existing:
                    cursor.execute('''
                        UPDATE crates
                        SET last_seen_lat = %s, last_seen_lng = %s, last_seen_time = %s, status = %s, battery = %s
                        WHERE id = %s
                    ''', (lat, lng, timestamp, status, battery, existing['id']))
                else:
                    cursor.execute('''
                        INSERT INTO crates (tag_mac, public_key, name, status, battery, last_seen_lat, last_seen_lng, last_seen_time, last_seen_address)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ''', (mac, pub_key, name, status, battery, lat, lng, timestamp, address))
                            
        conn.commit()
    except Exception as e:
        print("[HBB-K Sync] Genel Hata:", e)
    finally:
        conn.close()

# Rota Geçmişi (Frontend için CORS Proxy)
@app.route('/api/crates/<public_key>/route', methods=['GET'])
def get_crate_route(public_key):
    api_key = "hk_lZZ8t1fifpKfRuMsrcZb1AVSz161I_Ti"
    base_url = f"https://api.hbb-k.com/api/v1/devices/{public_key}/locations?limit=15"
    headers = {"X-API-KEY": api_key}
    
    try:
        response = requests.get(base_url, headers=headers, timeout=10)
        if response.status_code == 200:
            return jsonify(response.json())
        else:
            return jsonify({'error': 'Rota bilgisi alınamadı', 'status': response.status_code}), response.status_code
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Katalog Listeleme
@app.route('/api/catalog', methods=['GET'])
def get_catalog():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('SELECT * FROM parts_catalog')
        rows = cursor.fetchall()
        catalog = [dict(row) for row in rows]
        return jsonify(catalog)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# Excel Çıktısı (Export)
@app.route('/api/export', methods=['GET'])
def export_crates():
    sync_hbbk_data()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM crates')
    rows = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sandık Raporu"
    
    headers = ["ID", "Tag MAC", "Public Key", "PN", "Sandık Adı", "En (cm)", "Boy (cm)", "Yükseklik (cm)", "Durum", "Enlem", "Boylam", "Son Görülme Adresi", "Son Görülme Zamanı"]
    ws.append(headers)
    
    for row in rows:
        row_dict = dict(row)
        ws.append([
            row_dict['id'],
            row_dict['tag_mac'],
            row_dict['public_key'],
            row_dict.get('pn'),
            row_dict['name'],
            row_dict['width'],
            row_dict['length'],
            row_dict['height'],
            row_dict['status'],
            row_dict['last_seen_lat'],
            row_dict['last_seen_lng'],
            row_dict.get('last_seen_address'),
            row_dict['last_seen_time']
        ])
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Sandik_Raporu.xlsx'
    )

# Supabase ve Render için Keep-Alive Uç Noktası
@app.route('/api/keepalive', methods=['GET'])
def keepalive():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")  # Veritabanına ping atıp aktif tutar
        conn.close()
        return jsonify({
            'status': 'alive', 
            'message': 'Sunucu ve veritabanı uyanık durumda.',
            'time': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# AOG Listesini Excel'den Okuma
@app.route('/api/aog_list', methods=['GET'])
def get_aog_list():
    try:
        # Önce backend klasörü içinde (Render/Github için), yoksa bir üst klasörde ara (Lokal test için)
        local_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'AOG LIST.xlsx')
        parent_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'AOG LIST.xlsx')
        
        excel_path = local_path if os.path.exists(local_path) else parent_path
        
        if not os.path.exists(excel_path):
            return jsonify({'error': 'AOG LIST.xlsx bulunamadı. Lütfen Excel dosyasını sunucuya yüklediğinizden emin olun.'}), 404
            
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify([])
            
        headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
        aog_data = []
        for row in rows[1:]:
            if not row[0]: # PN boş ise satırı atla
                continue
            row_dict = {}
            for i, cell in enumerate(row):
                header = headers[i] if i < len(headers) else f'col_{i}'
                row_dict[header] = cell
            aog_data.append(row_dict)
            
        return jsonify(aog_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3000, debug=True)
