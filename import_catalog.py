import openpyxl
import psycopg2
import os
from database import get_db_connection

def import_catalog():
    excel_path = r'C:\Users\Elif\Downloads\AOG LIST.xlsx'
    
    print("Excel dosyası okunuyor...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"Excel dosyası okunamadı: {e}")
        return
    ws = wb.active
    
    # Hedef sütun indexlerini bul
    header = [cell.value for cell in ws[1]]
    
    try:
        idx_pn = header.index('PN')
        idx_desc = header.index('DESC')
        idx_length = header.index('Uzunluk')
        idx_width = header.index('Genişlik')
        idx_height = header.index('Yükseklik')
    except ValueError as e:
        print("Sütun başlıkları bulunamadı:", e)
        return

    conn = get_db_connection()
    cursor = conn.cursor()
    
    count = 0
    # 2. satırdan itibaren verileri oku
    for row in ws.iter_rows(min_row=2, values_only=True):
        pn = str(row[idx_pn]).strip() if row[idx_pn] else None
        if not pn or pn == 'None':
            continue
            
        desc = str(row[idx_desc]).strip() if row[idx_desc] else ""
        length = float(row[idx_length]) if row[idx_length] and str(row[idx_length]).replace('.', '', 1).isdigit() else 0.0
        width = float(row[idx_width]) if row[idx_width] and str(row[idx_width]).replace('.', '', 1).isdigit() else 0.0
        height = float(row[idx_height]) if row[idx_height] and str(row[idx_height]).replace('.', '', 1).isdigit() else 0.0
        
        try:
            cursor.execute(
                '''INSERT INTO parts_catalog (pn, description, width, length, height)
                   VALUES (%s, %s, %s, %s, %s)
                   ON CONFLICT (pn) DO UPDATE SET 
                   description = EXCLUDED.description, 
                   width = EXCLUDED.width, 
                   length = EXCLUDED.length, 
                   height = EXCLUDED.height''',
                (pn, desc, width, length, height)
            )
            count += 1
        except Exception as e:
            print(f"Hata ({pn}):", e)
            
    conn.commit()
    conn.close()
    print(f"Katalog başarıyla güncellendi. Toplam {count} parça eklendi.")

if __name__ == '__main__':
    import_catalog()
